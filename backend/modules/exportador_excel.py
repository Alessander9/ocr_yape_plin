"""
Exportador de resultados a Excel (.xlsx).
==========================================
Genera un archivo con 3 hojas: Resumen, Texto OCR, Auditoría.
"""

import json
from pathlib import Path
from typing import List
from datetime import datetime
from loguru import logger

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from modules.models import ResultadoOCR, NivelConfianza


# ── Colores ───────────────────────────────────────────────────────────────────
COLOR_HEADER_BG = "1E3A5F"    # Azul oscuro
COLOR_HEADER_FG = "FFFFFF"    # Blanco
COLOR_ALTA      = "C8E6C9"    # Verde claro
COLOR_MEDIA     = "FFF9C4"    # Amarillo claro
COLOR_BAJA      = "FFCDD2"    # Rojo claro
COLOR_GRIS      = "F5F5F5"    # Fila alternada


def _color_fila(nivel: NivelConfianza | None) -> str:
    if nivel == NivelConfianza.ALTA:
        return COLOR_ALTA
    elif nivel == NivelConfianza.MEDIA:
        return COLOR_MEDIA
    elif nivel == NivelConfianza.BAJA:
        return COLOR_BAJA
    return COLOR_GRIS


def _estilo_header(ws, fila: int, columnas: List[str]):
    """Aplica estilo al encabezado de una hoja."""
    fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
    font = Font(bold=True, color=COLOR_HEADER_FG, size=11)
    borde = Border(
        bottom=Side(style="medium", color="AAAAAA"),
    )
    for col_idx, col_name in enumerate(columnas, start=1):
        cell = ws.cell(row=fila, column=col_idx, value=col_name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde


def _ajustar_columnas(ws, min_width: int = 12, max_width: int = 45):
    """Ajusta el ancho de columnas automáticamente."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ancho = max(min_width, min(max_len + 3, max_width))
        ws.column_dimensions[col_letter].width = ancho


def _escribir_fila_resumen(ws, fila_num: int, r: ResultadoOCR):
    """Escribe una fila de datos en la hoja Resumen."""
    fila = [
        r.archivo,
        r.tipo_app.value if r.tipo_app else None,
        r.operacion_exitosa,
        r.monto,
        r.moneda,
        r.fecha,
        r.hora,
        r.nombre_emisor_o_pagador,
        r.nombre_receptor_o_destinatario,
        r.numero_celular_relacionado,
        r.banco_o_billetera,
        r.numero_operacion,
        r.descripcion,
        r.estado_operacion,
        r.confianza_global,
        ", ".join(r.campos_dudosos) if r.campos_dudosos else "",
        "; ".join(r.observaciones) if r.observaciones else "",
    ]

    color_bg = _color_fila(r.nivel_confianza)
    fill = PatternFill(fill_type="solid", fgColor=color_bg)

    for col_idx, valor in enumerate(fila, start=1):
        cell = ws.cell(row=fila_num, column=col_idx, value=valor)
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Formato específico por columna
        if col_idx == 4:  # monto
            cell.number_format = '#,##0.00'
        elif col_idx == 15:  # confianza
            cell.number_format = '0.00%'
            if isinstance(valor, float):
                cell.value = valor  # mantener como decimal para %


def exportar_excel(resultados: List[ResultadoOCR], ruta_salida: Path) -> Path:
    """
    Genera el archivo Excel completo con 3 hojas.
    Retorna la ruta del archivo generado.
    """
    wb = Workbook()

    # ── HOJA 1: Resumen ───────────────────────────────────────────────────────
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    columnas_resumen = [
        "Archivo", "Tipo App", "Op. Exitosa", "Monto", "Moneda",
        "Fecha", "Hora", "Emisor/Pagador", "Receptor/Destinatario",
        "Celular", "Banco/Billetera", "N° Operación", "Descripción",
        "Estado", "Confianza Global", "Campos Dudosos", "Observaciones",
    ]
    _estilo_header(ws_resumen, 1, columnas_resumen)
    ws_resumen.row_dimensions[1].height = 30

    for i, resultado in enumerate(resultados, start=2):
        _escribir_fila_resumen(ws_resumen, i, resultado)

    # Auto-filtro y filas congeladas
    ws_resumen.auto_filter.ref = ws_resumen.dimensions
    ws_resumen.freeze_panes = "A2"
    _ajustar_columnas(ws_resumen)

    # ── HOJA 2: Texto OCR ─────────────────────────────────────────────────────
    ws_texto = wb.create_sheet("Texto OCR")
    columnas_texto = ["Archivo", "Tipo App", "Texto Completo Detectado"]
    _estilo_header(ws_texto, 1, columnas_texto)

    for i, r in enumerate(resultados, start=2):
        fill = PatternFill(fill_type="solid", fgColor=COLOR_GRIS if i % 2 == 0 else "FFFFFF")
        for col_idx, valor in enumerate([
            r.archivo,
            r.tipo_app.value if r.tipo_app else "",
            r.texto_completo_detectado or "",
        ], start=1):
            cell = ws_texto.cell(row=i, column=col_idx, value=valor)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws_texto.freeze_panes = "A2"
    ws_texto.column_dimensions["A"].width = 30
    ws_texto.column_dimensions["B"].width = 15
    ws_texto.column_dimensions["C"].width = 80

    # ── HOJA 3: Auditoría ─────────────────────────────────────────────────────
    ws_audit = wb.create_sheet("Auditoría")
    columnas_audit = [
        "Archivo", "Calidad Imagen", "Motor OCR", "Tiempo (s)",
        "Confianza Global", "Nivel Confianza",
        "Conf. Monto", "Conf. Fecha", "Conf. Hora", "Conf. N°Op",
        "Conf. Emisor", "Conf. Receptor",
        "Posible Duplicado", "Hash Imagen",
    ]
    _estilo_header(ws_audit, 1, columnas_audit)

    for i, r in enumerate(resultados, start=2):
        cp = r.confianza_por_campo
        fila = [
            r.archivo,
            r.calidad_imagen_estimada,
            r.motor_ocr_primario,
            r.tiempo_procesamiento_seg,
            r.confianza_global,
            r.nivel_confianza.value if r.nivel_confianza else None,
            cp.monto if cp else None,
            cp.fecha if cp else None,
            cp.hora if cp else None,
            cp.numero_operacion if cp else None,
            cp.nombre_emisor_o_pagador if cp else None,
            cp.nombre_receptor_o_destinatario if cp else None,
            "Sí" if r.posible_duplicado else "No",
            r.hash_imagen,
        ]
        fill = PatternFill(fill_type="solid", fgColor=COLOR_GRIS if i % 2 == 0 else "FFFFFF")
        for col_idx, valor in enumerate(fila, start=1):
            cell = ws_audit.cell(row=i, column=col_idx, value=valor)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top")
            if col_idx in [5, 7, 8, 9, 10, 11, 12]:
                cell.number_format = '0.00%'

    ws_audit.freeze_panes = "A2"
    ws_audit.auto_filter.ref = ws_audit.dimensions
    _ajustar_columnas(ws_audit)

    # ── Metadatos del workbook ─────────────────────────────────────────────────
    wb.properties.title = "Reporte OCR Yape & Plin"
    wb.properties.creator = "Sistema OCR Yape & Plin v1.0"

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta_salida))
    logger.info(f"Excel exportado: {ruta_salida}")
    return ruta_salida
